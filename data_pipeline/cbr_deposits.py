"""Fetch household savings data from CBR Excel files.

Sources:
  1. 02_01_Funds_all.xlsx — aggregate deposits (with escrow breakdown)
  2. monetary_agg.xlsx — M2 structure (checking/term/FX breakdown)

Returns monthly volumes in billions RUB for each component
and the total (HH_SAVINGS_TOTAL ≈ 71 trln RUB as of early 2026).
"""

import io
import logging
from datetime import date, datetime

import pandas as pd
import requests

logger = logging.getLogger(__name__)

FUNDS_ALL_URL = (
    "https://cbr.ru/vfs/statistics/BankSector/Borrowings/02_01_Funds_all.xlsx"
)
MONETARY_AGG_URL = (
    "https://cbr.ru/vfs/statistics/credit_statistics/monetary_agg.xlsx"
)

HEADERS = {"User-Agent": "Mozilla/5.0"}


def _download_excel(url: str) -> bytes | None:
    """Download an Excel file from CBR, return raw bytes or None."""
    try:
        resp = requests.get(url, timeout=60, headers=HEADERS)
        resp.raise_for_status()
        return resp.content
    except requests.RequestException as e:
        logger.error("Failed to download %s: %s", url, e)
        return None


def _parse_funds_all(raw: bytes, date_from: date, date_to: date) -> pd.DataFrame:
    """Parse 02_01_Funds_all.xlsx 'итого' sheet.

    Extracts:
      Row 6: Вклады физ.лиц (с эскроу) → HH_SAVINGS_TOTAL
      Row 7: Вклады физ.лиц (без эскроу) → HH_DEPOSITS_NO_ESCROW
      Row 8: Средства на счетах эскроу → HH_ESCROW

    Values are in millions RUB in the file, converted to billions.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb["итого"]

    records = []
    row_map = {
        6: "HH_SAVINGS_TOTAL",
        7: "HH_DEPOSITS_NO_ESCROW",
        8: "HH_ESCROW",
    }

    for col in range(2, ws.max_column + 1):
        raw_date = ws.cell(2, col).value
        if raw_date is None:
            continue

        # Parse date (format: "01.02.2026" or datetime)
        if isinstance(raw_date, datetime):
            dt = raw_date.date()
        elif isinstance(raw_date, str):
            try:
                dt = datetime.strptime(raw_date.strip(), "%d.%m.%Y").date()
            except ValueError:
                continue
        else:
            continue

        if dt < date_from or dt > date_to:
            continue

        period = dt.strftime("%Y-%m-%d")

        for row_num, indicator in row_map.items():
            val = ws.cell(row_num, col).value
            if val is None or str(val).strip() in ("", "-"):
                continue
            try:
                numeric = float(str(val).replace(",", ".").replace(" ", ""))
            except ValueError:
                continue
            # Convert from millions to billions RUB
            records.append({
                "period_date": period,
                "indicator": indicator,
                "value": round(numeric / 1000, 2),
            })

    wb.close()
    logger.info("Parsed %d records from 02_01_Funds_all.xlsx", len(records))
    return pd.DataFrame(records)


def _parse_monetary_agg(raw: bytes, date_from: date, date_to: date) -> pd.DataFrame:
    """Parse monetary_agg.xlsx 'Денежные агрегаты' sheet.

    Extracts household deposit components:
      Row 7:  с начислением процентов → HH_SAVINGS_ACCOUNTS
      Row 8:  без начисления процентов → HH_CURRENT_ACCOUNTS
      Row 13: Другие депозиты домашних хозяйств → HH_TERM_DEPOSITS
      Row 18: Депозиты в иностранной валюте д/х → HH_FX_DEPOSITS

    Values are in billions RUB in the file (no conversion needed).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb["Денежные агрегаты"]

    records = []
    row_map = {
        7: "HH_SAVINGS_ACCOUNTS",
        8: "HH_CURRENT_ACCOUNTS",
        13: "HH_TERM_DEPOSITS",
        18: "HH_FX_DEPOSITS",
    }

    for col in range(2, ws.max_column + 1):
        raw_date = ws.cell(1, col).value
        if raw_date is None:
            continue

        if isinstance(raw_date, datetime):
            dt = raw_date.date()
        elif isinstance(raw_date, str):
            try:
                dt = datetime.strptime(raw_date.strip(), "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            continue

        if dt < date_from or dt > date_to:
            continue

        period = dt.strftime("%Y-%m-%d")

        for row_num, indicator in row_map.items():
            val = ws.cell(row_num, col).value
            if val is None or str(val).strip() in ("", "-"):
                continue
            try:
                numeric = float(str(val).replace(",", ".").replace(" ", ""))
            except ValueError:
                continue
            records.append({
                "period_date": period,
                "indicator": indicator,
                "value": round(numeric, 2),
            })

    wb.close()
    logger.info("Parsed %d records from monetary_agg.xlsx", len(records))
    return pd.DataFrame(records)


def fetch_household_savings(
    date_from: date = date(2010, 1, 1),
    date_to: date | None = None,
) -> pd.DataFrame:
    """Fetch all household savings components from CBR Excel files.

    Returns a DataFrame with columns: period_date, indicator, value.
    All values are in billions RUB.

    Indicators:
      HH_SAVINGS_TOTAL      — total deposits incl. escrow (from funds_all)
      HH_DEPOSITS_NO_ESCROW — deposits excl. escrow (from funds_all)
      HH_ESCROW             — escrow accounts (from funds_all)
      HH_TERM_DEPOSITS      — term deposits (from monetary_agg)
      HH_SAVINGS_ACCOUNTS   — interest-bearing checking (from monetary_agg)
      HH_CURRENT_ACCOUNTS   — non-interest checking (from monetary_agg)
      HH_FX_DEPOSITS        — foreign currency deposits (from monetary_agg)
    """
    if date_to is None:
        date_to = date.today()

    frames = []

    # Source 1: Aggregate deposits (with escrow breakdown)
    raw1 = _download_excel(FUNDS_ALL_URL)
    if raw1:
        try:
            df1 = _parse_funds_all(raw1, date_from, date_to)
            if not df1.empty:
                frames.append(df1)
        except Exception as e:
            logger.error("Failed to parse 02_01_Funds_all.xlsx: %s", e)

    # Source 2: M2 structure (checking/term/FX breakdown)
    raw2 = _download_excel(MONETARY_AGG_URL)
    if raw2:
        try:
            df2 = _parse_monetary_agg(raw2, date_from, date_to)
            if not df2.empty:
                frames.append(df2)
        except Exception as e:
            logger.error("Failed to parse monetary_agg.xlsx: %s", e)

    if not frames:
        logger.warning("No household savings data fetched from CBR")
        return pd.DataFrame()

    result = pd.concat(frames, ignore_index=True)

    # Log summary
    for ind in result["indicator"].unique():
        sub = result[result["indicator"] == ind]
        logger.info(
            "  %s: %d records, latest=%.1f bln RUB",
            ind, len(sub), sub["value"].iloc[-1] if not sub.empty else 0,
        )

    return result


# Keep backward-compatible alias
def fetch_household_deposits(
    date_from: date = date(2010, 1, 1),
    date_to: date | None = None,
) -> pd.DataFrame:
    """Backward-compatible wrapper. Returns only HH_SAVINGS_TOTAL rows."""
    df = fetch_household_savings(date_from, date_to)
    if df.empty:
        return df
    total = df[df["indicator"] == "HH_SAVINGS_TOTAL"].copy()
    if total.empty:
        return pd.DataFrame()
    total["indicator"] = "HH_DEPOSITS"
    return total
