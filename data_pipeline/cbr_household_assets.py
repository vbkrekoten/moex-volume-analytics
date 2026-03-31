"""Fetch full household financial assets from CBR quarterly balance sheet.

Source: https://cbr.ru/vfs/statistics/households/households_b.xlsx
Sheet: "Балансы" — quarterly data from Q1 2018.
All values in billions RUB.

Row map (based on CBR structure as of Q3 2025):
  R5:  АКТИВЫ (total)
  R6:  Наличная валюта (total cash)
  R7:  Наличная национальная валюта (cash RUB)
  R8:  Наличная иностранная валюта (cash FX)
  R9:  Депозиты (total deposits)
  R10: Переводные депозиты (demand deposits)
  R15: Другие депозиты (term deposits incl FX + precious metals)
  R18: Депозиты в банках-нерезидентах
  R21: Средства на брокерских счетах
  R24: Долговые ценные бумаги (bonds)
  R35: Займы (loans issued by households)
  R40: Акции и участие в капитале (equities total)
  R41: Котируемые акции резидентов
  R44: Некотируемые акции и прочее участие
  R45: Паи и акции инвестиционных фондов (резидентов)
  R48: Котируемые акции нерезидентов
  R55: Страховые и пенсионные резервы
  R59: Дебиторская задолженность
  R60: Средства на счетах эскроу
"""

import io
import logging
from datetime import date, datetime

import pandas as pd
import requests

logger = logging.getLogger(__name__)

HOUSEHOLDS_URL = "https://cbr.ru/vfs/statistics/households/households_b.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0"}

# Row number → indicator name (sheet "Балансы", 1-indexed)
ROW_MAP = {
    5: "HH_ASSETS_TOTAL",
    6: "HH_CASH_TOTAL",
    7: "HH_CASH_RUB",
    8: "HH_CASH_FX",
    9: "HH_DEPOSITS_TOTAL",
    10: "HH_DEPOSITS_DEMAND",
    15: "HH_DEPOSITS_TERM",
    18: "HH_DEPOSITS_NONRESIDENT",
    21: "HH_BROKERAGE",
    24: "HH_BONDS",
    35: "HH_LOANS_ISSUED",
    40: "HH_EQUITIES_TOTAL",
    41: "HH_EQUITIES_LISTED",
    44: "HH_EQUITIES_UNLISTED",
    45: "HH_FUNDS",
    48: "HH_EQUITIES_FOREIGN",
    55: "HH_INSURANCE_PENSION",
    59: "HH_RECEIVABLES",
    60: "HH_ESCROW_CBR",
}

# Russian labels for UI
INDICATOR_LABELS = {
    "HH_ASSETS_TOTAL": "Активы (всего)",
    "HH_CASH_TOTAL": "Наличные (всего)",
    "HH_CASH_RUB": "Наличные рубли",
    "HH_CASH_FX": "Наличная иностранная валюта",
    "HH_DEPOSITS_TOTAL": "Депозиты (всего)",
    "HH_DEPOSITS_DEMAND": "Переводные депозиты (текущие)",
    "HH_DEPOSITS_TERM": "Срочные депозиты (вкл. валютные)",
    "HH_DEPOSITS_NONRESIDENT": "Депозиты в банках-нерезидентах",
    "HH_BROKERAGE": "Брокерские счета",
    "HH_BONDS": "Долговые ценные бумаги (облигации)",
    "HH_LOANS_ISSUED": "Займы выданные",
    "HH_EQUITIES_TOTAL": "Акции и участие в капитале (всего)",
    "HH_EQUITIES_LISTED": "Котируемые акции (резиденты)",
    "HH_EQUITIES_UNLISTED": "Некотируемые акции и прочее участие",
    "HH_FUNDS": "Паи инвестиционных фондов",
    "HH_EQUITIES_FOREIGN": "Иностранные акции",
    "HH_INSURANCE_PENSION": "Страховые и пенсионные резервы",
    "HH_RECEIVABLES": "Дебиторская задолженность",
    "HH_ESCROW_CBR": "Счета эскроу",
}


def fetch_household_assets(
    date_from: date = date(2018, 1, 1),
    date_to: date | None = None,
) -> pd.DataFrame:
    """Download and parse CBR quarterly household balance sheet.

    Returns DataFrame with columns: period_date, indicator, value.
    Values are in billions RUB.
    """
    if date_to is None:
        date_to = date.today()

    try:
        resp = requests.get(HOUSEHOLDS_URL, timeout=60, headers=HEADERS)
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error("Failed to download households_b.xlsx: %s", e)
        return pd.DataFrame()

    return _parse_balance_sheet(resp.content, date_from, date_to)


def _parse_balance_sheet(
    raw: bytes, date_from: date, date_to: date,
) -> pd.DataFrame:
    """Parse the 'Балансы' sheet from households_b.xlsx."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb["Балансы"]

    records = []

    for col in range(2, ws.max_column + 1):
        raw_date = ws.cell(4, col).value  # Row 4 has dates
        if raw_date is None:
            continue

        if isinstance(raw_date, datetime):
            dt = raw_date.date()
        elif isinstance(raw_date, str):
            try:
                dt = datetime.strptime(raw_date.strip(), "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            continue

        if dt < date_from or dt > date_to:
            continue

        period = dt.strftime("%Y-%m-%d")

        for row_num, indicator in ROW_MAP.items():
            val = ws.cell(row_num, col).value
            if val is None or str(val).strip() in ("", "-"):
                continue
            try:
                numeric = float(str(val).replace(",", ".").replace(" ", ""))
            except ValueError:
                continue
            records.append({
                "period_date": period,
                "indicator": indicator,
                "value": round(numeric, 2),
            })

    wb.close()
    logger.info("Parsed %d records from households_b.xlsx", len(records))

    # Log summary
    df = pd.DataFrame(records)
    if not df.empty:
        for ind in sorted(df["indicator"].unique()):
            sub = df[df["indicator"] == ind]
            logger.info(
                "  %s: %d records, latest=%.1f bln",
                ind, len(sub), sub["value"].iloc[-1],
            )
    return df
